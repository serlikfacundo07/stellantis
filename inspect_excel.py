import pandas as pd
import openpyxl

files = [
    r"c:\Users\Administrator\Desktop\proyecto_tosi\excel\piezas.xlsm",
    r"c:\Users\Administrator\Desktop\proyecto_tosi\excel\Copia de SVP_Palomar_1805.xlsm NUEVO.xlsm"
]

for filepath in files:
    print(f"\n=== INSPECTING: {filepath} ===")
    try:
        # Load workbook to see sheet names
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print("Sheets:", wb.sheetnames)
        
        # Load the first sheet with pandas (first 10 rows, no header)
        sheet_name = wb.sheetnames[0]
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, nrows=20)
        print("First 20 rows:")
        for idx, row in df.iterrows():
            print(f"Row {idx}: {list(row.values[:15])}")
    except Exception as e:
        print(f"Error: {e}")
